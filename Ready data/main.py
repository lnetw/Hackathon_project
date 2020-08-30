import pandas as pd
import openpyxl
from pprint import pprint
all_data = {}

procentage_key = 'Процент людей вовлеченных в бизнес'
company_key = 'Общий оборот компании'
corup_key = 'Всего взяток'
vvp_key = 'Ввп'
health_key = 'Продолжительность жизни'
quality_key = 'Качество жизни'
internet_key = 'Пользуются выходом в сеть Интернет'
water_key = 'Горячее водоснабжение отсутствует'
gaz_key = 'Не имеют возможности пользоваться бытовым газом'
polyclinics_key = 'Удовлетворенность работой поликлиники'
hospitals_key = 'Комфортность условий пребывания в больнице'
stateOfHealth_key = 'Состояние своего здоровья'
alcohol_key = 'Не употребляют алкогольные напитки'
study_key = 'Удовлетворение качеством обучения'
work_key = 'Моральная удовлетворенность работой'
finance_expenses_key = 'Возможность справиться с неожиданными тратами'
finance_guest_key = 'Возможность приглашать гостей'
gasoline_key = 'Стоимость бензина АИ-92 на середину января 2020 г., руб./л'
rent_family_key = 'Доля семей, которые могут арендовать квартиру в 2019 г'
rent_profit_key = 'Семейный доход, необходимый для оплаты аренды и повседневных расходов, тыс. руб.'
payday_median_key = 'Медианная зарплата, тыс. руб'
payday_worker_key = 'Доля работающих с зарплатой выше 100 тыс. руб. в месяц, %'
zoh_key = 'Приверженность населения ЗОЖ'
social_spending_key = 'Доля социальных расходов в суммарных расходах консолидированного бюджета в 2019 году, %'
middle_class_key = 'Доля семей, относящихся к среднему классу, %'
unemployment_key = 'Уровень безработицы, %'
dtp_key = 'Число пострадавших в ДТП (погибших и раненых) на 100 тыс. человек в 2019 г.'
job_key = 'Изменение числа рабочих мест за три года, %'
crime_key = 'Колличесво преступлений за 2019 год'
eco_key_one = 'Рейтинг экологии городов в регионах по опросам'
eco_key_two = 'Индекс природо-охранный'
eco_key_three = 'Индекс промышленно охранный'
demography_key_one = 'Прирост (убыль) населения за 2017–2019 гг., %'
demography_key_two = 'Численность населения на 1 января 2020 г., тыс. чел.'
demography_key_three = 'Естественный прирост населения за 2017–2019 гг., тыс. чел.'
demography_key_four = 'Миграционный прирост населения за 2017–2019 гг., тыс. чел.'
demography_key_five = 'Коэфициент населения'


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_freedom.xlsx")
wb.active = 0
sheet = wb.active
for i in range(3,88):
    all_data.setdefault(str(sheet['B' + str(i)].value)[:-1], None)


for i in range(3,88):
    new_fitch = dict()
    new_fitch.setdefault(procentage_key, None)
    new_fitch[procentage_key] = float(str(sheet['C'+str(i)].value))
    new_fitch.setdefault(company_key, None)
    new_fitch[company_key] = float(str(sheet['E'+str(i)].value))
    reg_key = str(sheet['B' + str(i)].value)[:-1]
    all_data[reg_key] = new_fitch


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_goverment.xlsx", data_only=True)
wb.active = 0
sheet = wb.active
for i in range(2,87):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key in list(all_data.keys()):
        all_data[reg_key][corup_key] = int(str(sheet['D'+str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать!'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        all_data[reg_key][corup_key] = int(str(sheet['D' + str(i)].value))


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_gpd.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(4,98):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['B' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][vvp_key] = float(str(sheet['B' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['B' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][vvp_key] = float(str(sheet['B' + str(i)].value))



wb = openpyxl.reader.excel.load_workbook(filename="Table_data_health.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(5,99):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['B' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][health_key] = float(str(sheet['B' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['B' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][vvp_key] = float(str(sheet['B' + str(i)].value))


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_quality.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(2,99):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['B' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][quality_key] = float(str(sheet['B' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['B' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][quality_key] = float(str(sheet['B' + str(i)].value))


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_internet.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(10,102):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['C' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][internet_key] = float(str(sheet['C' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['C' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][internet_key] = float(str(sheet['C' + str(i)].value))


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_water.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(10,102):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['E' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][water_key] = float(str(sheet['E' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['E' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][water_key] = float(str(sheet['E' + str(i)].value))


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_gaz.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(9,101):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['F' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][gaz_key] = float(str(sheet['F' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['F' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][gaz_key] = float(str(sheet['F' + str(i)].value))


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_polyclinics.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(10,102):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['D' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][polyclinics_key] = float(str(sheet['D' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['D' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][polyclinics_key] = float(str(sheet['D' + str(i)].value))



wb = openpyxl.reader.excel.load_workbook(filename="Table_data_hospitals.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(10,102):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['D' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][hospitals_key] = float(str(sheet['D' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['D' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][hospitals_key] = float(str(sheet['D' + str(i)].value))



wb = openpyxl.reader.excel.load_workbook(filename="Table_data_state_health.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(9,101):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['C' + str(i)].value) == 'None' or str(sheet['D' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][stateOfHealth_key] = float(str(sheet['D' + str(i)].value)) + float(str(sheet['C' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['C' + str(i)].value) == 'None' or str(sheet['D' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][stateOfHealth_key] = float(str(sheet['D' + str(i)].value)) + float(str(sheet['C' + str(i)].value))



wb = openpyxl.reader.excel.load_workbook(filename="Table_data_alcohol.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(10,102):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['I' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][alcohol_key] = float(str(sheet['I' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['C' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][alcohol_key] = float(str(sheet['I' + str(i)].value))


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_study.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(10,102):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['D' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][study_key] = float(str(sheet['D' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['D' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][study_key] = float(str(sheet['D' + str(i)].value))



wb = openpyxl.reader.excel.load_workbook(filename="Table_data_work.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(10,102):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['AE' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][work_key] = float(str(sheet['AE' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['AE' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][work_key] = float(str(sheet['AE' + str(i)].value))


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_finance.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(10,102):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['D' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][finance_expenses_key] = float(str(sheet['D' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['D' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][finance_expenses_key] = float(str(sheet['D' + str(i)].value))

for i in range(10,102):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['J' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][finance_guest_key] = float(str(sheet['J' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['J' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][finance_guest_key] = float(str(sheet['J' + str(i)].value))


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_gasoline.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['D' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][gasoline_key] = float(str(sheet['D' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['D' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][gasoline_key] = float(str(sheet['D' + str(i)].value))
pprint(all_data)


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_rent.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['C' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][rent_family_key] = float(str(sheet['C' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['C' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][rent_family_key] = float(str(sheet['C' + str(i)].value))

for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['D' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][rent_profit_key] = float(str(sheet['D' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['D' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][rent_profit_key] = float(str(sheet['D' + str(i)].value))



wb = openpyxl.reader.excel.load_workbook(filename="Table_data_payday.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['E' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][payday_median_key] = float(str(sheet['E' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['E' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][payday_median_key] = float(str(sheet['E' + str(i)].value))


for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['C' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][payday_worker_key] = float(str(sheet['C' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['C' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][payday_worker_key] = float(str(sheet['C' + str(i)].value))


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_zoh.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['C' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][zoh_key] = float(str(sheet['C' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['C' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][zoh_key] = float(str(sheet['C' + str(i)].value))


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_social_spending.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['E' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][social_spending_key] = float(str(sheet['E' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['E' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][social_spending_key] = float(str(sheet['E' + str(i)].value))



wb = openpyxl.reader.excel.load_workbook(filename="Table_data_middle_class.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['C' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][middle_class_key] = float(str(sheet['C' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['C' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][middle_class_key] = float(str(sheet['C' + str(i)].value))


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_unemployment.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['C' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][unemployment_key] = float(str(sheet['C' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['C' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][unemployment_key] = float(str(sheet['C' + str(i)].value))


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_dtp.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['E' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][dtp_key] = int(str(sheet['E' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['E' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][dtp_key] = int(str(sheet['E' + str(i)].value))



wb = openpyxl.reader.excel.load_workbook(filename="Table_data_job.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['E' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][job_key] = float(str(sheet['E' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['E' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][job_key] = float(str(sheet['E' + str(i)].value))


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_crime.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(2,97):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['B' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][crime_key] = int(str(sheet['B' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['B' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][crime_key] = int(str(sheet['B' + str(i)].value))
pprint(all_data)

wb = openpyxl.reader.excel.load_workbook(filename="Table_data_eco_socaity.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(2,87):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['B' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][eco_key_one] = float(str(sheet['B' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['B' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][eco_key_one] = float(str(sheet['B' + str(i)].value))


for i in range(2,87):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['C' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][eco_key_two] = float(str(sheet['C' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['C' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][eco_key_two] = float(str(sheet['C' + str(i)].value))


for i in range(2,87):
    reg_key = str(sheet['A' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['D' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][eco_key_three] = float(str(sheet['D' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['D' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][eco_key_three] = float(str(sheet['D' + str(i)].value))
pprint(all_data)


wb = openpyxl.reader.excel.load_workbook(filename="Table_data_demography.xlsx", data_only=True)
wb.active = 0
sheet = wb.active

for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['C' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][demography_key_one] = float(str(sheet['C' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['C' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][demography_key_one] = float(str(sheet['C' + str(i)].value))

for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['D' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][demography_key_two] = float(str(sheet['D' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['D' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][demography_key_two] = float(str(sheet['D' + str(i)].value))


for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['E' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][demography_key_three] = float(str(sheet['E' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['E' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][demography_key_three] = float(str(sheet['E' + str(i)].value))



for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['F' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][demography_key_four] = float(str(sheet['F' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['F' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][demography_key_four] = float(str(sheet['F' + str(i)].value))


for i in range(2,87):
    reg_key = str(sheet['B' + str(i)].value)
    if reg_key == 'None' or reg_key is None:
        continue
    if reg_key in list(all_data.keys()):
        if str(sheet['G' + str(i)].value) == 'None':
            continue
        else:
            all_data[reg_key][demography_key_five] = float(str(sheet['G' + str(i)].value))
    else:
        print('{} не совпадает со словарем, требуется преименовать! Введите cont, если нужно пропустить'.format(reg_key))
        reg_key = input('Нужное имя региона: ')
        if reg_key == 'cont':
            continue
        if reg_key in list(all_data.keys()):
            if str(sheet['G' + str(i)].value) == 'None':
                continue
            else:
                all_data[reg_key][demography_key_five] = float(str(sheet['G' + str(i)].value))


table = pd.DataFrame(data=None, index=None, columns=['Регион',
                                                     'Процент людей вовлеченных в бизнес',
                                                     'Общий оборот компании',
                                                     'Всего взяток',
                                                     'Ввп',
                                                     'Продолжительность жизни',
                                                     'Качество жизни',
                                                     internet_key,
                                                     water_key,
                                                     gaz_key,
                                                     polyclinics_key,
                                                     hospitals_key,
                                                     stateOfHealth_key,
                                                     alcohol_key,
                                                     study_key,
                                                     work_key,
                                                     finance_expenses_key,
                                                     finance_guest_key,
                                                     gasoline_key,
                                                     rent_family_key,
                                                     rent_profit_key,
                                                     payday_median_key,
                                                     payday_worker_key,
                                                     zoh_key,
                                                     social_spending_key,
                                                     middle_class_key,
                                                     unemployment_key,
                                                     dtp_key,
                                                     job_key,
                                                     crime_key,
                                                     eco_key_one,
                                                     eco_key_two,
                                                     eco_key_three,
                                                     demography_key_one,
                                                     demography_key_two,
                                                     demography_key_three,
                                                     demography_key_four,
                                                     demography_key_five],
                     dtype=None,
                     copy=False)
dict_append = {}
for reg in all_data:
    dict_append['Регион'] = reg
    dict_append[procentage_key] = all_data[reg][procentage_key]
    dict_append[company_key] = all_data[reg][company_key]
    dict_append[corup_key] = all_data[reg][corup_key]
    dict_append[vvp_key] = all_data[reg][vvp_key]
    dict_append[health_key] = all_data[reg][health_key]
    dict_append[quality_key] = all_data[reg][quality_key]
    dict_append[internet_key] = all_data[reg][internet_key]
    dict_append[water_key] = all_data[reg][water_key]
    dict_append[gaz_key] = all_data[reg][gaz_key]
    dict_append[polyclinics_key] = all_data[reg][polyclinics_key]
    dict_append[hospitals_key] = all_data[reg][hospitals_key]
    dict_append[stateOfHealth_key] = all_data[reg][stateOfHealth_key]
    dict_append[alcohol_key] = all_data[reg][alcohol_key]
    dict_append[study_key] = all_data[reg][study_key]
    dict_append[work_key] = all_data[reg][work_key]
    dict_append[finance_expenses_key] = all_data[reg][finance_expenses_key]
    dict_append[finance_guest_key] = all_data[reg][finance_guest_key]
    dict_append[gasoline_key] = all_data[reg][gasoline_key]
    dict_append[rent_family_key] = all_data[reg][rent_family_key]
    dict_append[rent_profit_key] = all_data[reg][rent_profit_key]
    dict_append[payday_median_key] = all_data[reg][payday_median_key]
    dict_append[payday_worker_key] = all_data[reg][payday_worker_key]
    dict_append[zoh_key] = all_data[reg][zoh_key]
    dict_append[social_spending_key] = all_data[reg][social_spending_key]
    dict_append[middle_class_key] = all_data[reg][middle_class_key]
    dict_append[unemployment_key] = all_data[reg][unemployment_key]
    dict_append[dtp_key] = all_data[reg][dtp_key]
    dict_append[job_key] = all_data[reg][job_key]
    dict_append[crime_key] = all_data[reg][crime_key]
    dict_append[eco_key_one] = all_data[reg][eco_key_one]
    dict_append[eco_key_two] = all_data[reg][eco_key_two]
    dict_append[eco_key_three] = all_data[reg][eco_key_three]
    dict_append[demography_key_one] = all_data[reg][demography_key_one]
    dict_append[demography_key_two] = all_data[reg][demography_key_two]
    dict_append[demography_key_three] = all_data[reg][demography_key_three]
    dict_append[demography_key_four] = all_data[reg][demography_key_four]
    dict_append[demography_key_five] = all_data[reg][demography_key_five]
    table = table.append(dict_append, ignore_index=True)
    dict_append.clear()
table.to_csv('table_output.csv', index=False, header=True)
table.to_excel('table_excel.xlsx', index=False, header=True)
pprint(table)

